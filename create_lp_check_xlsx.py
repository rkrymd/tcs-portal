import subprocess
import sys

# Install openpyxl
subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Common styles
header_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
header_font = Font(bold=True, size=11)
edit_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
red_font = Font(color="FF0000", size=10)
normal_font = Font(size=10)
wrap_align = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

col_widths = [18, 28, 55, 55, 32]
headers = ["セクション", "項目", "現状の文言", "変更後（記入欄）", "備考"]


def setup_sheet(ws, data, warnings=None):
    """Setup a sheet with headers, data, and formatting."""
    if warnings is None:
        warnings = {}

    # Headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Column widths
    for col_idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value else "")
            cell.alignment = wrap_align
            cell.border = thin_border
            cell.font = normal_font

            # Yellow background for "変更後" column
            if col_idx == 4:
                cell.fill = edit_fill

        # Red font for warning notes in 備考 column
        if row_idx in warnings:
            cell = ws.cell(row=row_idx, column=5)
            cell.font = red_font

    # Freeze panes (below header)
    ws.freeze_panes = "A2"


# ============================================================
# Sheet 1: Pattern A（白）
# ============================================================
ws1 = wb.active
ws1.title = "Pattern A（白）"

data_a = [
    # Hero
    ["Hero", "バッジ", "🎬 TikTokクリエイター向け", "", ""],
    ["Hero", "見出し", "好きな商品を紹介して報酬を得よう", "", ""],
    ["Hero", "サブテキスト", "TCSは、TikTokクリエイターと人気ブランドをつなぐアフィリエイトプログラム。専用ポータルで案件管理、特別料率で高報酬。", "", ""],
    ["Hero", "CTAボタン", "LINEで無料登録", "", ""],
    ["Hero", "注釈", "最短30秒で登録完了 ・ 費用は一切かかりません", "", ""],
    # 実績
    ["実績", "ラベル", "Results", "", ""],
    ["実績", "タイトル", "実績で選ばれています", "", ""],
    ["実績", "数字1", "120+ / 提携クリエイター", "", ""],
    ["実績", "数字2", "40+ / 取扱ブランド・商品", "", ""],
    ["実績", "数字3", "20% / 最大報酬料率", "", ""],
    ["実績", "数字4", "¥59万 / 月間TOP報酬額", "", ""],
    # 選ばれる理由
    ["選ばれる理由", "タイトル", "TCSが選ばれる理由", "", ""],
    ["選ばれる理由", "サブ", "一般的なアフィリエイトとは違う、クリエイターファーストの仕組み。", "", ""],
    ["選ばれる理由", "特徴1", "💰 特別料率で高報酬 / TikTok Shopの公開料率よりも高い特別料率を提供。", "", ""],
    ["選ばれる理由", "特徴2", "📱 専用ポータルで案件管理 / 案件の検索・応募・進捗管理がすべてポータルで完結。", "", ""],
    ["選ばれる理由", "特徴3", "🎁 サンプル無料提供 / 紹介する商品のサンプルを無料でお届け。", "", ""],
    ["選ばれる理由", "特徴4", "⚡ 審査なしで即参加OK / 一部の案件は審査不要で即参加可能。", "", ""],
    ["選ばれる理由", "特徴5", "🎬 動画もLIVEもOK / 投稿スタイルは自由。", "", ""],
    ["選ばれる理由", "特徴6", "🤝 専任サポート付き / LINEですぐに相談可能。", "", ""],
    # ポータル
    ["ポータル", "タイトル", "専用ポータルで案件を一括管理", "", ""],
    ["ポータル", "説明", "登録するとクリエイター専用ポータルにアクセスできます。", "", ""],
    ["ポータル", "商品例1", "スキンケアセット A / 報酬率 15%", "", ""],
    ["ポータル", "商品例2", "オーガニックコーヒー / 報酬率 12%", "", ""],
    ["ポータル", "商品例3", "リップグロス 新色 / 報酬率 20%", "", ""],
    # 始め方
    ["始め方", "タイトル", "始めるのはとても簡単", "", ""],
    ["始め方", "Step1", "LINEで無料登録 / 30秒で登録完了", "", ""],
    ["始め方", "Step2", "案件を選んで応募", "", ""],
    ["始め方", "Step3", "商品を紹介 / サンプルが届いたら自由に紹介", "", ""],
    ["始め方", "Step4", "報酬GET / 売れた分だけ報酬発生", "", ""],
    # クリエイターの声
    ["クリエイターの声", "タイトル", "クリエイターの声", "", ""],
    ["クリエイターの声", "声1", "「ポータルが使いやすくて案件選びが楽」/ 美容系 3.2万人", "", "※プレースホルダー。実際の声に差し替え推奨"],
    ["クリエイターの声", "声2", "「料率が高い。月20万以上稼げてます」/ ライフスタイル系 5.8万人", "", "※プレースホルダー"],
    ["クリエイターの声", "声3", "「サンプルが届くので良い商品だけ紹介できる」/ フード系 1.5万人", "", "※プレースホルダー"],
    # FAQ
    ["FAQ", "Q1", "費用はかかりますか？ → 完全無料", "", ""],
    ["FAQ", "Q2", "フォロワー数の条件は？ → 特に下限なし", "", ""],
    ["FAQ", "Q3", "対象ジャンルは？ → 幅広いジャンル", "", ""],
    ["FAQ", "Q4", "報酬の受け取りは？ → TikTok Shop経由", "", ""],
    ["FAQ", "Q5", "LIVE配信のみ？ → 動画投稿でもOK", "", ""],
    # CTA
    ["CTA", "見出し", "あなたのTikTokで報酬を得よう", "", ""],
    ["CTA", "CTAボタン", "LINEで無料登録する", "", ""],
]

# Rows with red warning text in 備考 (1-indexed from data, +1 for header = row_idx)
warnings_a = {32: True, 33: True, 34: True}  # クリエイターの声 rows

setup_sheet(ws1, data_a, warnings_a)

# ============================================================
# Sheet 2: Pattern B（ダーク）
# ============================================================
ws2 = wb.create_sheet("Pattern B（ダーク）")

data_b = [
    # Hero
    ["Hero", "アイキャッチ", "TikTok Creator Partner Program", "", ""],
    ["Hero", "見出し", "TikTokで稼げるを、本気で。", "", ""],
    ["Hero", "サブ", "人気ブランド × 特別料率 × 専用ツール", "", ""],
    ["Hero", "数字1", "¥59万 / 月間TOP報酬", "", ""],
    ["Hero", "数字2", "20% / 最大報酬率", "", ""],
    ["Hero", "数字3", "40+ / 取扱商品", "", ""],
    ["Hero", "CTAボタン", "今すぐ参加する", "", ""],
    ["Hero", "注釈", "30秒で登録完了 ・ 完全無料", "", ""],
    # 比較
    ["比較", "タイトル", "普通のアフィリエイトと何が違う？", "", ""],
    ["比較", "一般アフィリ", "公開料率のみ/案件自分で探す/サポートなし/サンプル自腹/管理手動", "", ""],
    ["比較", "TCSパートナー", "特別料率/ポータルで案件提案/LINE専任サポート/サンプル無料/自動管理", "", ""],
    # 報酬
    ["報酬", "タイトル", "報酬率が全然違う", "", ""],
    ["報酬", "バー1", "5% / 公開料率", "", ""],
    ["報酬", "バー2", "15% / TCS特別料率", "", ""],
    ["報酬", "バー3", "20% / 招待限定", "", ""],
    # 始め方
    ["始め方", "タイトル", "4ステップで始められる", "", ""],
    ["始め方", "Step1", "LINE登録 / 30秒で完了", "", ""],
    ["始め方", "Step2", "案件を選ぶ", "", ""],
    ["始め方", "Step3", "紹介する / 動画でもLIVEでもOK", "", ""],
    ["始め方", "Step4", "報酬GET", "", ""],
    # ポータル
    ["ポータル", "タイトル", "クリエイター専用ポータル", "", ""],
    ["ポータル", "機能", "案件検索/進捗管理/LINE通知/お気に入り", "", ""],
    # ブランド
    ["ブランド", "タイトル", "人気ブランドと提携", "", ""],
    ["ブランド", "一覧", "こうじや/のむシリカ/UCC/ROSABLU/Vioteras/KINS/リクセル/イングリウッド", "", "※実ブランド名。公開前に許可確認必要"],
    # CTA
    ["CTA", "見出し", "始めるなら、今。", "", ""],
    ["CTA", "CTAボタン", "今すぐ参加する", "", ""],
]

warnings_b = {25: True}  # ブランド一覧 row

setup_sheet(ws2, data_b, warnings_b)

# ============================================================
# Sheet 3: 共通メモ
# ============================================================
ws3 = wb.create_sheet("共通メモ")

memo_data = [
    ["項目", "内容", "ステータス"],
    ["LINE登録URL", "https://lin.ee/XXXXX", "→ 実URLに差替え必要"],
    ["Voices（クリエイターの声）", "プレースホルダー", "→ 実際の声に差替え推奨"],
    ["Brands（ブランド名）", "実名使用", "→ 許可確認必要"],
    ["Preview URL - Pattern A", "https://rkrymd.github.io/tcs-portal/lp_a.html", ""],
    ["Preview URL - Pattern B", "https://rkrymd.github.io/tcs-portal/lp_b.html", ""],
]

for row_idx, row_data in enumerate(memo_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = wrap_align
        cell.border = thin_border
        if row_idx == 1:
            cell.font = header_font
            cell.fill = header_fill
        elif col_idx == 3 and value.startswith("→"):
            cell.font = red_font
        else:
            cell.font = normal_font

ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 55
ws3.column_dimensions["C"].width = 30
ws3.freeze_panes = "A2"

# Save
output_path = r"C:\claude-projects\tcs-portal\LP文言チェックシート.xlsx"
wb.save(output_path)
print(f"Created: {output_path}")
