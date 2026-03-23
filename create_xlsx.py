import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

header_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
header_font = Font(bold=True, size=11)
yellow_fill = PatternFill(start_color='FFFDE7', end_color='FFFDE7', fill_type='solid')
red_font = Font(color='CC0000', size=11)
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)
wrap_align = Alignment(wrap_text=True, vertical='top')

columns = ['セクション', '項目', '現状の文言', '変更後の文言（記入欄）', '備考']
col_widths = [18, 28, 55, 55, 32]

def setup_sheet(ws, data_rows):
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical='center')
        cell.border = thin_border
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = 'A2'
    for row_idx, row_data in enumerate(data_rows, 2):
        section, item, text, biko = row_data
        ws.cell(row=row_idx, column=1, value=section).alignment = wrap_align
        ws.cell(row=row_idx, column=2, value=item).alignment = wrap_align
        ws.cell(row=row_idx, column=3, value=text).alignment = wrap_align
        c4 = ws.cell(row=row_idx, column=4, value='')
        c4.fill = yellow_fill
        c4.alignment = wrap_align
        c5 = ws.cell(row=row_idx, column=5, value=biko)
        c5.alignment = wrap_align
        if biko:
            c5.font = red_font
        for ci in range(1, 6):
            ws.cell(row=row_idx, column=ci).border = thin_border

# Sheet 1: Pattern A
ws1 = wb.active
ws1.title = 'Pattern A（白）'
pa = [
    ('Hero', 'バッジ', '🎬 TikTokクリエイター向け', ''),
    ('Hero', '見出し（h1）', '好きな商品を紹介して報酬を得よう', ''),
    ('Hero', 'サブテキスト', 'TCSは、TikTokクリエイターと人気ブランドをつなぐアフィリエイトプログラム。専用ポータルで案件管理、特別料率で高報酬。', ''),
    ('Hero', 'CTAボタン', 'LINEで無料登録', ''),
    ('Hero', '注釈', '最短30秒で登録完了 ・ 費用は一切かかりません', ''),
    ('実績', 'セクションラベル', 'Results', ''),
    ('実績', 'セクションタイトル', '実績で選ばれています', ''),
    ('実績', '数字1', '120+ / 提携クリエイター', ''),
    ('実績', '数字2', '40+ / 取扱ブランド・商品', ''),
    ('実績', '数字3', '20% / 最大報酬料率', ''),
    ('実績', '数字4', '¥59万 / 月間TOP報酬額', ''),
    ('選ばれる理由', 'セクションタイトル', 'TCSが選ばれる理由', ''),
    ('選ばれる理由', 'サブテキスト', '一般的なアフィリエイトとは違う、クリエイターファーストの仕組み。', ''),
    ('選ばれる理由', '特徴1 見出し', '💰 特別料率で高報酬', ''),
    ('選ばれる理由', '特徴1 説明', 'TikTok Shopの公開料率よりも高い特別料率を提供。売れた分だけしっかり報酬が入ります。', ''),
    ('選ばれる理由', '特徴2 見出し', '📱 専用ポータルで案件管理', ''),
    ('選ばれる理由', '特徴2 説明', '案件の検索・応募・進捗管理がすべてポータルで完結。LINEでリアルタイム通知も届きます。', ''),
    ('選ばれる理由', '特徴3 見出し', '🎁 サンプル無料提供', ''),
    ('選ばれる理由', '特徴3 説明', '紹介する商品のサンプルを無料でお届け。実際に使ってからレビューできます。', ''),
    ('選ばれる理由', '特徴4 見出し', '⚡ 審査なしで即参加OK', ''),
    ('選ばれる理由', '特徴4 説明', '一部の案件は審査不要で即参加可能。登録したその日から報酬を得られます。', ''),
    ('選ばれる理由', '特徴5 見出し', '🎬 動画もLIVEもOK', ''),
    ('選ばれる理由', '特徴5 説明', '投稿スタイルは自由。レビュー動画でもLIVE配信でも、あなたの得意な方法で紹介できます。', ''),
    ('選ばれる理由', '特徴6 見出し', '🤝 専任サポート付き', ''),
    ('選ばれる理由', '特徴6 説明', '困ったことがあればLINEですぐに相談可能。オリエンシートや参考動画も充実しています。', ''),
    ('ポータル', 'セクションタイトル', '専用ポータルで案件を一括管理', ''),
    ('ポータル', '説明', '登録するとクリエイター専用ポータルにアクセスできます。案件の検索・応募・進捗確認がすべてここで完結。', ''),
    ('ポータル', '商品例1', 'スキンケアセット A / 報酬率 15%', ''),
    ('ポータル', '商品例2', 'オーガニックコーヒー / 報酬率 12%', ''),
    ('ポータル', '商品例3', 'リップグロス 新色 / 報酬率 20%', ''),
    ('始め方', 'セクションタイトル', '始めるのはとても簡単', ''),
    ('始め方', 'Step1', 'LINEで無料登録 / LINEアカウントで30秒で登録完了。TikTok IDとジャンルを入力するだけ。', ''),
    ('始め方', 'Step2', '案件を選んで応募 / 専用ポータルで気になる商品を見つけて応募。即参加OKの案件もあります。', ''),
    ('始め方', 'Step3', '商品を紹介 / サンプルが届いたら、動画やLIVEで自由に紹介。スタイルは完全にお任せ。', ''),
    ('始め方', 'Step4', '報酬GET / 売れた分だけ報酬が発生。成果はポータルとLINEでリアルタイム確認。', ''),
    ('クリエイターの声', '声1', '「ポータルが使いやすくて案件選びが楽。LINEで通知が来るから見逃さない」/ 美容系クリエイター ・ フォロワー3.2万人', '※プレースホルダー。実際の声に差し替え推奨'),
    ('クリエイターの声', '声2', '「他のアフィリエイトより料率が高い。LIVE配信で月20万以上稼げてます」/ ライフスタイル系 ・ フォロワー5.8万人', '※プレースホルダー'),
    ('クリエイターの声', '声3', '「サンプルが届くので本当に良いと思った商品だけ紹介できるのが嬉しい」/ フード系クリエイター ・ フォロワー1.5万人', '※プレースホルダー'),
    ('FAQ', 'Q1', '費用はかかりますか？ → 一切かかりません。登録も利用も完全無料です。', ''),
    ('FAQ', 'Q2', 'フォロワー数の条件はありますか？ → 特に下限は設けていません。コンテンツの質や熱意を重視しています。', ''),
    ('FAQ', 'Q3', 'どんなジャンルのクリエイターが対象ですか？ → 美容・フード・ライフスタイル・ガジェットなど幅広いジャンルの案件があります。', ''),
    ('FAQ', 'Q4', '報酬はいつ・どうやって受け取れますか？ → TikTok Shopのアフィリエイト報酬として、TikTokの仕組みを通じて支払われます。', ''),
    ('FAQ', 'Q5', 'LIVE配信じゃないとダメですか？ → 動画投稿でもOKです。あなたの得意なスタイルで紹介してください。', ''),
    ('CTA', '見出し', 'あなたのTikTokで報酬を得よう', ''),
    ('CTA', 'サブテキスト', '登録は無料・最短30秒で完了。今すぐ専用ポータルで案件をチェック。', ''),
    ('CTA', 'CTAボタン', 'LINEで無料登録する', ''),
]
setup_sheet(ws1, pa)

# Sheet 2: Pattern B
ws2 = wb.create_sheet('Pattern B（ダーク）')
pb = [
    ('Hero', 'アイキャッチ', 'TikTok Creator Partner Program', ''),
    ('Hero', '見出し', 'TikTokで稼げるを、本気で。', ''),
    ('Hero', 'サブ見出し', '人気ブランド × 特別料率 × 専用ツール', ''),
    ('Hero', '数字1', '¥59万 / 月間TOP報酬', ''),
    ('Hero', '数字2', '20% / 最大報酬率', ''),
    ('Hero', '数字3', '40+ / 取扱商品', ''),
    ('Hero', 'CTAボタン', '今すぐ参加する', ''),
    ('比較', 'セクションタイトル', '普通のアフィリエイトと何が違う？', ''),
    ('比較', '一般アフィリ', '公開料率のみ（低い）/ 案件は自分で探す / サポートなし / サンプルは自腹 / 進捗管理は手動', ''),
    ('比較', 'TCSパートナー', '特別料率で高報酬 / 専用ポータルで案件提案 / LINE専任サポート / サンプル無料提供 / ワークフロー自動管理', ''),
    ('報酬', 'セクションタイトル', '報酬率が全然違う', ''),
    ('報酬', '説明', 'TikTok Shopの公開料率と比較して、TCSパートナーには特別料率を提供。招待案件ならさらに高報酬。', ''),
    ('報酬', 'バー表示', '5% 公開料率 / 15% TCS特別料率 / 20% 招待限定', ''),
    ('始め方', 'Step1', 'LINE登録 / 30秒で完了。TikTok IDを入力するだけ', ''),
    ('始め方', 'Step2', '案件を選ぶ / ポータルで気になる商品を見つけて応募', ''),
    ('始め方', 'Step3', '紹介する / 動画でもLIVEでもOK。スタイルは自由', ''),
    ('始め方', 'Step4', '報酬GET / 売れた分だけ報酬発生。成果はリアルタイム確認', ''),
    ('ポータル', 'セクションタイトル', 'クリエイター専用ポータル', ''),
    ('ポータル', '説明', '案件の検索・応募・進捗管理がすべて1つのポータルで完結。他にはないクリエイター体験。', ''),
    ('ポータル', '機能一覧', '案件検索&フィルター / 応募・進捗ステータス管理 / LINEリアルタイム通知 / お気に入り&レコメンド', ''),
    ('ブランド', 'セクションタイトル', '人気ブランドと提携', ''),
    ('ブランド', 'ブランド一覧', 'こうじや / のむシリカ / UCC / ROSABLU / Vioteras / KINS / リクセル / イングリウッド', '※実ブランド名。公開前に許可確認必要'),
    ('CTA', '見出し', '始めるなら、今。', ''),
    ('CTA', 'サブテキスト', '登録は完全無料。あなたのTikTokが収益に変わる。', ''),
    ('CTA', 'CTAボタン', '今すぐ参加する', ''),
]
setup_sheet(ws2, pb)

# Sheet 3: 共通メモ
ws3 = wb.create_sheet('共通メモ')
ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 80
ws3.cell(row=1, column=1, value='LP共通の注意事項').font = Font(bold=True, size=14)
ws3.cell(row=3, column=1, value='要対応')
ws3.cell(row=3, column=2, value='LINE登録URL: https://lin.ee/XXXXX → 実際のURLに差し替え必要')
ws3.cell(row=4, column=1, value='要対応')
ws3.cell(row=4, column=2, value='Voices: プレースホルダー → 実クリエイターの声に差し替え推奨')
ws3.cell(row=5, column=1, value='要確認')
ws3.cell(row=5, column=2, value='Brands: Pattern Bで実名使用 → 公開前に許可確認が必要')
ws3.cell(row=7, column=1, value='Preview URLs').font = Font(bold=True, size=12)
ws3.cell(row=8, column=1, value='Pattern A')
ws3.cell(row=8, column=2, value='https://ripples-inc.github.io/tcs-portal/lp_a.html')
ws3.cell(row=9, column=1, value='Pattern B')
ws3.cell(row=9, column=2, value='https://ripples-inc.github.io/tcs-portal/lp_b.html')

output_path = r'C:\claude-projects\tcs-portal\LP文言チェックシート.xlsx'
wb.save(output_path)
fsize = os.path.getsize(output_path)
print(f'SUCCESS: {output_path}')
print(f'Sheets: {wb.sheetnames}')
print(f'Pattern A: {ws1.max_row} rows, Pattern B: {ws2.max_row} rows, 共通メモ: {ws3.max_row} rows')
print(f'File size: {fsize} bytes')
